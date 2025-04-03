# # Fixing the code for better structure and naming convention

# import pandas as pd

# # Load the Excel file from the provided path
# file_path_xlsx = 'Technology.xlsx'
# xls_new = pd.ExcelFile(file_path_xlsx)

# # Load the specific sheet 'CriteriaSurveyReport'
# df = pd.read_excel(file_path_xlsx, sheet_name='CriteriaSurveyReport')

# # Step 1: Clean the dataframe by dropping rows and columns that are completely empty
# df_cleaned = df.dropna(how='all').dropna(axis=1, how='all')

# # Step 2: Reset the index after cleaning for better readability
# df_cleaned = df_cleaned.reset_index(drop=True)

# # Step 3: Review the number of columns and rename accordingly
# # Since the column count is 8, we'll only rename the first 8 columns.
# df_cleaned.columns = df_cleaned.columns[:8]  # Adjust to rename only first 8

# # Now renaming the first 8 columns as per structure.
# df_cleaned.columns = ["No.", "Col2", "Question Text", "Response", "Col5", "Survey Score", "Col7", "Comment"]

# # Step 4: Select only relevant columns for further processing
# df_final_cleaned = df_cleaned[["No.", "Question Text", "Response", "Comment"]]

# # Step 5: Drop rows where both "No." and "Question Text" are NaN as these are critical fields
# df_final_cleaned = df_final_cleaned.dropna(subset=["No.", "Question Text"])

# # Print the cleaned dataframe for the user
# print(df_final_cleaned)



# import cv2
# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image as openpyxlImage
# from PIL import Image as PILImage
# import numpy as np
# import os

# # Load the Excel file
# file_path = 'Technology.xlsx'
# wb = load_workbook(file_path)
# ws = wb['CriteriaSurveyReport']  # Adjust to your sheet name

# # Directory to store extracted images
# extracted_image_dir = 'extracted_images'
# os.makedirs(extracted_image_dir, exist_ok=True)

# # Function to save and load the extracted image from Excel
# def save_extracted_image(img, img_name):
#     img_path = os.path.join(extracted_image_dir, img_name)
#     img.save(img_path)
#     return img_path

# # Extract all images in the worksheet
# extracted_image_paths = []
# for idx, image in enumerate(ws._images):
#     img = PILImage.open(image.ref)
#     img_name = f"extracted_image_{idx}.png"
#     img_path = save_extracted_image(img, img_name)
#     extracted_image_paths.append(img_path)

# # Load the image to compare (checkbox image)
# checkbox_image_path = 'image_9.png'
# checkbox_image = cv2.imread(checkbox_image_path, 0)  # Load as grayscale

# # Function to compare two images
# def compare_images(img1_path, img2_path):
#     # Load both images in grayscale
#     img1 = cv2.imread(img1_path, 0)
#     img2 = cv2.imread(img2_path, 0)

#     # Check if sizes match
#     if img1.shape != img2.shape:
#         return False

#     # Compare images using cv2.absdiff (absolute difference)
#     difference = cv2.absdiff(img1, img2)
    
#     # If no difference, the images match
#     if cv2.countNonZero(difference) == 0:
#         return True
#     return False

# # Compare the checkbox image with the extracted images
# for img_path in extracted_image_paths:
#     if compare_images(checkbox_image_path, img_path):
#         print(f"Match found with image: {img_path}")
#     else:
#         print(f"No match for image: {img_path}")


import cv2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as openpyxlImage
from PIL import Image as PILImage
import numpy as np
import os

# Load the Excel file (your data Excel file)
file_path = 'Technology.xlsx'
wb = load_workbook(file_path)
ws = wb['CriteriaSurveyReport']  # Adjust this to your sheet name

# Directory to store extracted images
extracted_image_dir = 'extracted_images'
os.makedirs(extracted_image_dir, exist_ok=True)

# Load your main Excel data into a DataFrame
df = pd.read_excel(file_path, sheet_name='CriteriaSurveyReport')

# Step 1: Clean the DataFrame by dropping rows and columns that are completely empty
df_cleaned = df.dropna(how='all').dropna(axis=1, how='all')
df_cleaned = df_cleaned.reset_index(drop=True)

# Review the number of columns and rename accordingly (based on your previous code)
df_cleaned.columns = df_cleaned.columns[:8]
df_cleaned.columns = ["No.", "Col2", "Question Text", "Response", "Col5", "Survey Score", "Col7", "Comment"]

# Select only relevant columns
df_final_cleaned = df_cleaned[["No.", "Question Text", "Response", "Comment"]]

# Step 2: Save and load the extracted image from Excel
def save_extracted_image(img, img_name):
    img_path = os.path.join(extracted_image_dir, img_name)
    img.save(img_path)
    return img_path

# Extract all images in the worksheet
extracted_image_paths = []
image_to_cell_map = {}  # To map extracted images to their corresponding cell locations
for idx, image in enumerate(ws._images):
    img = PILImage.open(image.ref)
    img_name = f"extracted_image_{idx}.png"
    img_path = save_extracted_image(img, img_name)
    extracted_image_paths.append(img_path)
    image_to_cell_map[img_path] = image.anchor  # Map the image to the cell location

# Load the checkbox image to compare
checkbox_image_path = 'image_9.png'
checkbox_image = cv2.imread(checkbox_image_path, 0)  # Load as grayscale

# Function to compare two images
def compare_images(img1_path, img2_path):
    img1 = cv2.imread(img1_path, 0)
    img2 = cv2.imread(img2_path, 0)
    if img1.shape != img2.shape:
        return False
    difference = cv2.absdiff(img1, img2)
    if cv2.countNonZero(difference) == 0:
        return True
    return False

# Step 3: Compare images and update the DataFrame's Response column
for img_path in extracted_image_paths:
    if compare_images(checkbox_image_path, img_path):
        # Find the corresponding cell for the matched image
        cell_location = image_to_cell_map[img_path]

        # Extract row and column information from OneCellAnchor
        row_index = cell_location._from.row  # Get the row index of the image anchor
        col_index = cell_location._from.col  # Get the column index of the image anchor
        
        # Adjust row index if necessary (Excel's 0-based row index)
        # Set the response in the DataFrame for the matched row
        df_final_cleaned.at[row_index, 'Response'] = 'Matched Checkbox Image'

# Step 4: Display the final DataFrame
# import ace_tools as tools; tools.display_dataframe_to_user(name="Final DataFrame with Response Updated", dataframe=df_final_cleaned)

# Alternatively, if running locally, print the DataFrame:
print(df_final_cleaned)
